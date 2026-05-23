#!/usr/bin/env python3
"""
build_catalog_db.py — promote Instrument Workshop Master v3.xlsx to SQLite.

Brainstorm Tier 3 #8: the Master Catalog xlsx becomes a *view* over the db,
not the source of truth.

Reads the named tables in the workbook, normalizes column names, infers
column types, and writes a single SQLite file. Idempotent — rebuilds from
scratch each run.

Usage:
    python3 scripts/build_catalog_db.py "/path/to/Instrument Workshop Master v3.xlsx"
    python3 scripts/build_catalog_db.py "/path/to/...xlsx" --output ./catalog.sqlite
    python3 scripts/build_catalog_db.py "/path/to/...xlsx" --dry-run
"""

import argparse
import re
import sqlite3
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl",
          file=sys.stderr)
    sys.exit(1)


def normalize_col(name: str) -> str:
    """ 'Instrument ID' -> 'instrument_id'; 'Time Hours' -> 'time_hours' """
    s = re.sub(r"[^A-Za-z0-9]+", "_", str(name).strip().lower())
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "col"


def normalize_table(name: str) -> str:
    return normalize_col(name)


def infer_type(values):
    """Look at non-empty values and pick INTEGER, REAL, or TEXT."""
    seen_int, seen_float, seen_other = False, False, False
    for v in values:
        if v is None or v == "":
            continue
        if isinstance(v, bool):
            seen_other = True
        elif isinstance(v, int):
            seen_int = True
        elif isinstance(v, float):
            if v.is_integer() and not seen_float:
                seen_int = True
            else:
                seen_float = True
        else:
            seen_other = True
            break
    if seen_other:
        return "TEXT"
    if seen_float:
        return "REAL"
    if seen_int:
        return "INTEGER"
    return "TEXT"


def coerce(value, sql_type):
    if value is None or value == "":
        return None
    if sql_type == "TEXT":
        return str(value)
    if sql_type == "INTEGER":
        try:
            return int(value)
        except (TypeError, ValueError):
            return None
    if sql_type == "REAL":
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    return str(value)


def read_table(ws, ref):
    """Read a table by its A1:X266-style ref. Returns (headers, rows)."""
    cells = list(ws[ref])
    if not cells:
        return [], []
    headers = [str(c.value).strip() if c.value is not None else f"col_{i}"
               for i, c in enumerate(cells[0])]
    rows = []
    for row in cells[1:]:
        # Skip rows where every cell is empty
        if all(c.value is None or c.value == "" for c in row):
            continue
        rows.append([c.value for c in row])
    return headers, rows


def find_named_tables(wb):
    """Yield (sheet_name, table_name, ref) for every defined Excel table."""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for tbl in getattr(ws, "tables", {}).values():
            yield sheet, tbl.name, tbl.ref


def primary_key_for(headers):
    """Heuristic: the first column that ends in '_id' is the PK; else first col."""
    for h in headers:
        n = normalize_col(h)
        if n.endswith("_id"):
            return n
    return normalize_col(headers[0])


INDEX_PLAN = [
    # (sql_table_name, [(index_name, column)])
    ("master_catalog", [
        ("idx_master_family", "family"),
        ("idx_master_status", "build_status"),
        ("idx_master_priority", "priority"),
    ]),
    ("production_log", [
        ("idx_production_instrument", "instrument_id"),
        ("idx_production_status", "status"),
    ]),
    ("doe_studies", [
        ("idx_doe_instrument", "instrument_id"),
    ]),
    ("cad_cnc_library", [
        ("idx_cad_readiness", "cnc_readiness"),
    ]),
]


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("workbook", help="Path to Instrument Workshop Master v3.xlsx")
    ap.add_argument("--output", default="./catalog.sqlite",
                    help="Output SQLite path (default: ./catalog.sqlite)")
    ap.add_argument("--dry-run", action="store_true",
                    help="Print schema and row counts; don't write the .sqlite")
    args = ap.parse_args()

    wb_path = Path(args.workbook)
    if not wb_path.exists():
        print(f"workbook not found: {wb_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(wb_path, data_only=True)

    # Build a plan: list of (table_name, headers, rows) per Excel table.
    plan = []
    for sheet, tbl_name, ref in find_named_tables(wb):
        ws = wb[sheet]
        headers, rows = read_table(ws, ref)
        if not headers:
            continue
        # Friendly table name: prefer sheet name (Master Catalog -> master_catalog)
        sql_table = normalize_table(sheet)
        plan.append({
            "sheet": sheet,
            "tbl_name": tbl_name,
            "ref": ref,
            "sql_table": sql_table,
            "headers": headers,
            "rows": rows,
        })

    if not plan:
        print("No named Excel tables found in workbook.", file=sys.stderr)
        sys.exit(1)

    # Print plan to stdout
    print(f"workbook: {wb_path}")
    print(f"tables: {len(plan)}")
    for entry in plan:
        sql_cols = []
        col_types = []
        for i, h in enumerate(entry["headers"]):
            col_name = normalize_col(h)
            col_values = [r[i] for r in entry["rows"]] if entry["rows"] else []
            col_type = infer_type(col_values)
            sql_cols.append(col_name)
            col_types.append(col_type)
        entry["sql_cols"] = sql_cols
        entry["col_types"] = col_types
        entry["pk"] = primary_key_for(entry["headers"])
        print(f"  {entry['sheet']!r:<24} -> {entry['sql_table']:<24}  "
              f"{len(entry['headers'])} cols, {len(entry['rows'])} rows  "
              f"(pk={entry['pk']})")

    if args.dry_run:
        print("\n--dry-run: not writing the .sqlite")
        return 0

    out = Path(args.output)
    if out.exists():
        out.unlink()
    conn = sqlite3.connect(out)
    cur = conn.cursor()

    for entry in plan:
        cols_def = []
        for col, t in zip(entry["sql_cols"], entry["col_types"]):
            if col == entry["pk"]:
                cols_def.append(f"{col} {t} PRIMARY KEY")
            else:
                cols_def.append(f"{col} {t}")
        ddl = f"CREATE TABLE {entry['sql_table']} ({', '.join(cols_def)});"
        cur.execute(ddl)

        placeholders = ", ".join("?" for _ in entry["sql_cols"])
        ins = (f"INSERT OR REPLACE INTO {entry['sql_table']} "
               f"({', '.join(entry['sql_cols'])}) VALUES ({placeholders})")
        for row in entry["rows"]:
            coerced = []
            for v, t in zip(row, entry["col_types"]):
                coerced.append(coerce(v, t))
            # Skip rows whose primary-key value is empty
            pk_idx = entry["sql_cols"].index(entry["pk"])
            if coerced[pk_idx] in (None, ""):
                continue
            cur.execute(ins, coerced)

    # Indices
    for sql_table, idx_list in INDEX_PLAN:
        # Only create if the table actually exists in plan
        if not any(e["sql_table"] == sql_table for e in plan):
            continue
        target = next(e for e in plan if e["sql_table"] == sql_table)
        for idx_name, idx_col in idx_list:
            if idx_col not in target["sql_cols"]:
                continue
            cur.execute(f"CREATE INDEX {idx_name} "
                        f"ON {sql_table}({idx_col});")

    conn.commit()
    conn.close()

    print(f"\nwrote {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
