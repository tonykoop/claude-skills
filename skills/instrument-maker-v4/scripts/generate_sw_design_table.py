#!/usr/bin/env python3
"""
generate_sw_design_table.py — emit a SolidWorks design table .xlsx from a
family-spec.csv.

The output is a single-sheet workbook in the format SolidWorks expects:

    Row 1: free-text title (SW ignores)
    Row 2: empty | $DESCRIPTION | $COLOR | $VALUE@<global>@Equations | ...
    Row 3+: one row per family-spec member, formula values like =18in / =0.5in

Tony then in SolidWorks runs:
    Insert > Tables > Design Table > From Existing File
and points it at the produced xlsx. SW imports the configurations.

Usage:
    python3 scripts/generate_sw_design_table.py \\
        ./build-packets/<slug> \\
        --output ./build-packets/<slug>/cad/sw-design-table.xlsx \\
        --part-name TNG-001_TongueDrum

Optional:
    --global-prefix dim_      prefix family-spec column names with this
    --color 15000804          $COLOR value (default RGB-encoded)
    --dry-run                 print the proposed structure; don't write

Human-readable headers are normalized before promotion:
    Scale Length (in)  -> scale_length_in
    Body Thickness mm  -> body_thickness_mm
    Variant            -> member_id
"""

import argparse
import csv
import re
import sys
from pathlib import Path


# Columns from family-spec.csv that should NOT become SW globals.
DENY_LIST_COLS = {
    "member_id", "id", "target_note", "scale_label", "wood_species",
    "notes", "k_constant", "k2_correction"
}

# Suffixes whose columns are *outputs / acoustic targets*, not CAD globals.
# We deny-list these too because they shouldn't drive SW geometry.
ACOUSTIC_SUFFIXES = ("_hz", "_freq", "_cents", "_ratio", "_db", "_pct")

# Heuristics for picking up unit from column name.
INCH_SUFFIXES = ("_in", "_inch", "_inches")  # require leading underscore
MM_SUFFIXES = ("_mm",)
ANGLE_SUFFIXES = ("_deg", "_rad", "_angle")
COUNT_SUFFIXES = ("_count", "_n", "_num", "_qty")

UNIT_SUFFIXES = {
    "in": "in",
    "inch": "in",
    "inches": "in",
    '"': "in",
    "mm": "mm",
    "millimeter": "mm",
    "millimeters": "mm",
    "deg": "deg",
    "degree": "deg",
    "degrees": "deg",
    "rad": "rad",
    "radian": "rad",
    "radians": "rad",
    "hz": "hz",
}

HEADER_ALIASES = {
    "variant": "member_id",
    "configuration": "member_id",
    "config": "member_id",
    "member": "member_id",
    "member_name": "member_id",
    "note": "target_note",
    "target": "target_note",
    "target_note": "target_note",
    "description": "notes",
}


def is_numeric(value: str) -> bool:
    if value is None or value == "":
        return False
    try:
        float(str(value).strip().replace(",", ""))
        return True
    except (TypeError, ValueError):
        return False


def normalize_header(header: str) -> str:
    """Map human-readable family-spec headers to SW-safe snake_case.

    Examples:
      `Scale Length (in)` -> `scale_length_in`
      `Body Thickness [mm]` -> `body_thickness_mm`
      `Variant` -> `member_id`

    The script still denies acoustic target columns later, so
    `Target Hz` becomes `target_hz` and is intentionally not promoted.
    """
    raw = (header or "").strip()
    lower = raw.lower().strip()

    unit_suffix = ""
    unit_match = re.search(r"[\(\[]\s*([^\)\]]+?)\s*[\)\]]\s*$", lower)
    if unit_match:
        unit_key = unit_match.group(1).strip().replace(".", "")
        unit_suffix = UNIT_SUFFIXES.get(unit_key, "")
        lower = lower[: unit_match.start()].strip()

    lower = lower.replace("%", " pct ")
    lower = lower.replace("#", " count ")
    lower = lower.replace("&", " and ")
    lower = lower.replace("@", " at ")
    lower = re.sub(r"[^a-z0-9]+", "_", lower).strip("_")
    lower = re.sub(r"_+", "_", lower)
    if not lower:
        lower = "column"
    lower = HEADER_ALIASES.get(lower, lower)

    if unit_suffix and not lower.endswith(f"_{unit_suffix}"):
        lower = f"{lower}_{unit_suffix}"
    return lower


def normalize_rows(raw_rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], dict[str, str]]:
    """Return rows keyed by normalized headers plus a raw->normalized map."""
    if not raw_rows:
        return [], {}
    raw_headers = list(raw_rows[0].keys())
    used: dict[str, int] = {}
    header_map: dict[str, str] = {}
    for raw in raw_headers:
        normalized = normalize_header(raw)
        count = used.get(normalized, 0)
        used[normalized] = count + 1
        if count:
            normalized = f"{normalized}_{count + 1}"
        header_map[raw] = normalized

    rows: list[dict[str, str]] = []
    for raw_row in raw_rows:
        rows.append({header_map[k]: (v or "") for k, v in raw_row.items()})
    return rows, header_map


def load_openpyxl():
    """Import openpyxl only when writing xlsx; --dry-run stays dependency-light."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("openpyxl is required to write .xlsx files. Install with: pip install openpyxl",
              file=sys.stderr)
        return None
    return Workbook, Font, PatternFill


def detect_unit(column_name: str) -> str:
    n = column_name.lower()
    for suf in MM_SUFFIXES:
        if n.endswith(suf):
            return "mm"
    for suf in INCH_SUFFIXES:
        if n.endswith(suf):
            return "in"
    for suf in ANGLE_SUFFIXES:
        if n.endswith(suf):
            return "deg" if n.endswith(("_deg", "_angle")) else "rad"
    for suf in COUNT_SUFFIXES:
        if n.endswith(suf):
            return ""
    # Conservative: if no suffix hint, leave unitless. SW will treat as a
    # raw number, which is safer than mis-attributing units like "in" to
    # a value that's actually a frequency or a ratio.
    return ""


def is_promotable_global(column_name: str, sample_values) -> bool:
    if column_name in DENY_LIST_COLS:
        return False
    if not column_name:
        return False
    # Must be snake_case-ish: only [a-zA-Z0-9_], starting with a letter
    if not re.fullmatch(r"[a-zA-Z][a-zA-Z0-9_]*", column_name):
        return False
    # Reject acoustic-output columns — they aren't CAD globals.
    cl = column_name.lower()
    for suf in ACOUSTIC_SUFFIXES:
        if cl.endswith(suf):
            return False
    # Must have at least one numeric sample
    return any(is_numeric(v) for v in sample_values if v not in ("", None))


def format_value(value: str, unit: str) -> str:
    s = str(value).strip()
    if not s:
        return ""
    # If already starts with =, pass through
    if s.startswith("="):
        return s
    if not is_numeric(s):
        return f'"{s}"'
    if unit:
        return f"={s}{unit}"
    return f"={s}"


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("packet", type=Path,
                    help="Path to a build-packet directory (must contain "
                         "family-spec.csv).")
    ap.add_argument("--output", type=Path,
                    help="Output xlsx path (default: <packet>/cad/"
                         "sw-design-table.xlsx)")
    ap.add_argument("--part-name", default="Instrument",
                    help="SolidWorks part name (used in row 1 title)")
    ap.add_argument("--global-prefix", default="",
                    help="Optional prefix to prepend to each global name "
                         "(e.g., 'dim_')")
    ap.add_argument("--color", default="15000804",
                    help="$COLOR value for each row (default 15000804)")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    fs_path = args.packet / "family-spec.csv"
    if not fs_path.exists():
        print(f"family-spec.csv not found at {fs_path}", file=sys.stderr)
        return 1

    with fs_path.open(newline="", encoding="utf-8-sig") as f:
        raw_rows = list(csv.DictReader(f))
    rows, header_map = normalize_rows(raw_rows)

    if not rows:
        print(f"family-spec.csv is empty: {fs_path}", file=sys.stderr)
        return 1

    headers = list(rows[0].keys())
    promotable = []
    for h in headers:
        samples = [row.get(h) for row in rows]
        if is_promotable_global(h, samples):
            promotable.append(h)
    if not promotable:
        print(f"no promotable global columns in {fs_path}.\n"
              f"All headers were rejected. Headers: {headers}",
              file=sys.stderr)
        return 1

    print(f"family-spec: {len(rows)} rows, {len(headers)} columns")
    if any(raw != normalized for raw, normalized in header_map.items()):
        print("header map:")
        for raw, normalized in header_map.items():
            print(f"  {raw!r} -> {normalized!r}")
    print(f"promoting {len(promotable)} columns to SW globals: "
          f"{promotable}")

    out = args.output or (args.packet / "cad" / "sw-design-table.xlsx")

    if args.dry_run:
        print(f"\n--dry-run — would write {out}")
        print(f"  row 1: Design Table for: {args.part_name}")
        header2 = [""] + ["$DESCRIPTION", "$COLOR"] + [
            f"$VALUE@{args.global_prefix}{h}@Equations"
            for h in promotable]
        print(f"  row 2 ({len(header2)} cols): {header2}")
        for i, row in enumerate(rows):
            mid = (row.get("member_id") or row.get("target_note")
                   or f"M-{i+1}")
            desc = row.get("notes", mid)
            r = [mid, desc, args.color]
            for h in promotable:
                r.append(format_value(row.get(h, ""), detect_unit(h)))
            print(f"  row {i+3}: {r}")
        return 0

    out.parent.mkdir(parents=True, exist_ok=True)
    openpyxl_api = load_openpyxl()
    if openpyxl_api is None:
        return 1
    Workbook, Font, PatternFill = openpyxl_api

    wb = Workbook()
    ws = wb.active
    ws.title = "DesignTable"

    # Row 1: title
    ws.cell(row=1, column=1, value=f"Design Table for: {args.part_name}")
    ws.cell(row=1, column=1).font = Font(bold=True)

    # Row 2: SW header
    ws.cell(row=2, column=1, value="")
    ws.cell(row=2, column=2, value="$DESCRIPTION")
    ws.cell(row=2, column=3, value="$COLOR")
    for j, h in enumerate(promotable):
        ws.cell(row=2, column=4 + j,
                value=f"$VALUE@{args.global_prefix}{h}@Equations")
    # Bold header
    for c in range(1, 4 + len(promotable)):
        ws.cell(row=2, column=c).font = Font(bold=True)
        ws.cell(row=2, column=c).fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Configuration rows
    for i, row in enumerate(rows):
        mid = (row.get("member_id") or row.get("target_note")
               or f"M-{i+1}")
        desc = row.get("notes", mid)
        r_idx = 3 + i
        ws.cell(row=r_idx, column=1, value=mid)
        ws.cell(row=r_idx, column=2, value=desc)
        ws.cell(row=r_idx, column=3, value=args.color)
        for j, h in enumerate(promotable):
            val = row.get(h, "")
            unit = detect_unit(h)
            ws.cell(row=r_idx, column=4 + j,
                    value=format_value(val, unit))

    # Auto-width columns
    n_cols = 3 + len(promotable)
    n_rows = 2 + len(rows)
    for c in range(1, n_cols + 1):
        col_letter = ws.cell(row=2, column=c).column_letter
        widths = [len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, n_rows + 1)]
        max_len = max(widths) if widths else 10
        ws.column_dimensions[col_letter].width = min(40, max(10, max_len + 2))

    wb.save(out)
    print("wrote " + str(out))
    print("  " + str(len(rows)) + " configurations, " + str(n_cols) + " columns")
    print("\nNext steps in SolidWorks:")
    print("  1. Open " + args.part_name + ".SLDPRT")
    print("  2. Insert > Tables > Design Table > From Existing File")
    print("  3. Browse to: " + str(out))
    print("  4. SW will import " + str(len(rows)) + " configurations")
    print("  5. Verify globals exist in Tools > Equations:")
    for h in promotable[:6]:
        print("       " + args.global_prefix + h)
    if len(promotable) > 6:
        print("       ... (" + str(len(promotable) - 6) + " more)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
