#!/usr/bin/env python3
"""
ingest_dimension_csv.py — validate SolidWorks dimensions vs Excel design table.

Reads the CSV emitted by Extract_Dimensions.swp and diffs it against the
Excel design workbook (Master_Inputs sheet by default). Reports matches,
SW-only, Excel-only, and mismatches with a tolerance.

Usage:
    python3 scripts/ingest_dimension_csv.py \\
        --csv /path/to/dimensions.csv \\
        --workbook /path/to/Drone-Flutes-Design.xlsx \\
        --design-sheet Master_Inputs

Optional:
    --config <name>           restrict to one assembly configuration
    --tolerance-percent 0.5   default ±0.5%
    --report findings.md      structured report path
    --write-validation packet/validation.csv  append findings to a packet
    --dry-run                 print summary; don't write report or validation
"""

import argparse
import csv
import re
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl",
          file=sys.stderr)
    sys.exit(1)


# Map a SW dimension full name like
#   "drum_len_in_sketch@SK_OUTER_BOX@TNG-000_MasterLayout.Part"
# to a clean global-style identifier "drum_len_in" by:
#   1. Taking everything before the first '@'
#   2. Stripping a trailing '_sketch' if present
NAME_AT_PART = re.compile(r"^([^@]+)@")


def normalize_dim_name(full_name: str) -> str:
    m = NAME_AT_PART.match(full_name.strip())
    base = m.group(1) if m else full_name.strip()
    if base.endswith("_sketch"):
        base = base[: -len("_sketch")]
    return base


def parse_equation_lhs(eq: str) -> str:
    """ '"drum_len_in"= 18in' → 'drum_len_in' """
    m = re.match(r'\s*"([^"]+)"\s*=', eq)
    return m.group(1) if m else ""


def parse_equation_rhs(eq: str):
    """ '"drum_len_in"= 18in' → ('18', 'in') """
    m = re.search(r'=\s*([0-9.\-+eE]+)\s*([A-Za-z]*)', eq)
    if not m:
        return None, None
    val_str, unit = m.group(1), m.group(2)
    try:
        return float(val_str), unit.lower()
    except ValueError:
        return None, unit.lower()


def in_to_in(value, unit):
    if value is None:
        return None
    u = (unit or "").lower()
    if u in ("", "in", "inch", "inches"):
        return value
    if u == "mm":
        return value / 25.4
    if u == "cm":
        return value / 2.54
    if u == "m":
        return value * 39.3700787
    return value  # unknown unit — return as-is


def read_sw_csv(path: Path, only_config: str | None = None):
    """Returns dict: { config_name: { global_name: value_in } }"""
    out: dict[str, dict[str, float]] = {}
    with path.open(newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            cfg = row.get("AssemblyConfigName", "").strip()
            if only_config and cfg != only_config:
                continue
            if cfg not in out:
                out[cfg] = {}
            # Two extraction strategies: from the equation column (preferred
            # for global vars) and from DimFullName (for feature dimensions).
            eq = row.get("EquationOrComment", "")
            is_global = row.get("IsGlobalVar", "").upper() == "TRUE"
            value_in = None
            try:
                value_in = float(row.get("Value_in", ""))
            except (TypeError, ValueError):
                value_in = None
            if is_global and eq:
                name = parse_equation_lhs(eq)
                if name:
                    if value_in is None:
                        v, u = parse_equation_rhs(eq)
                        value_in = in_to_in(v, u)
                    out[cfg][name] = value_in
            else:
                full = row.get("DimFullName", "")
                if full:
                    name = normalize_dim_name(full)
                    if name and value_in is not None:
                        # Don't overwrite a global-derived value with a
                        # feature-derived value; globals are authoritative.
                        out[cfg].setdefault(name, value_in)
    return out


def read_excel_master_inputs(workbook: Path, sheet_name: str):
    """Return dict { global_name: value_in } from the Excel side.

    Heuristic:
      - Look for a column named 'Variable' / 'Name' (any case).
      - The next non-empty column is the value.
      - If the value cell has units in a separate 'Unit' column, normalize
        to inches.
    """
    wb = load_workbook(workbook, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"sheet not found in {workbook}: {sheet_name} "
                         f"(available: {wb.sheetnames})")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    out: dict[str, float] = {}

    # Find header row containing 'Variable' or 'Name'
    header_row_idx = None
    name_col = None
    val_col = None
    unit_col = None
    for i, row in enumerate(rows):
        for j, cell in enumerate(row or []):
            if isinstance(cell, str):
                cl = cell.strip().lower()
                if cl in ("variable", "name", "global", "parameter"):
                    header_row_idx = i
                    name_col = j
                    break
        if header_row_idx is not None:
            # Find value and unit columns to the right
            row = rows[header_row_idx]
            for j, cell in enumerate(row or []):
                if isinstance(cell, str):
                    cl = cell.strip().lower()
                    if cl in ("value", "value (in)", "value_in") and val_col is None:
                        val_col = j
                    elif cl in ("unit", "units"):
                        unit_col = j
            if val_col is None and name_col is not None:
                val_col = name_col + 1
            break

    if header_row_idx is None:
        # Fallback: scan all rows; treat any (string, number) pair as
        # (name, value). Only useful for very simple sheets.
        for row in rows:
            if not row or len(row) < 2:
                continue
            n, v = row[0], row[1]
            if isinstance(n, str) and isinstance(v, (int, float)):
                out[n.strip()] = float(v)
        return out

    for i in range(header_row_idx + 1, len(rows)):
        row = rows[i]
        if not row or name_col >= len(row):
            continue
        name = row[name_col]
        if not isinstance(name, str) or not name.strip():
            continue
        if val_col >= len(row):
            continue
        val = row[val_col]
        if not isinstance(val, (int, float)):
            continue
        unit = ""
        if unit_col is not None and unit_col < len(row):
            u = row[unit_col]
            if isinstance(u, str):
                unit = u.strip()
        out[name.strip()] = in_to_in(float(val), unit)
    return out


def diff(sw: dict[str, float], excel: dict[str, float],
         tolerance_pct: float):
    matches, sw_only, excel_only, mismatches = [], [], [], []
    sw_names = set(sw.keys())
    excel_names = set(excel.keys())
    for name in sorted(sw_names | excel_names):
        sw_val = sw.get(name)
        ex_val = excel.get(name)
        if sw_val is None and ex_val is None:
            continue
        if sw_val is None:
            excel_only.append((name, ex_val))
            continue
        if ex_val is None:
            sw_only.append((name, sw_val))
            continue
        # Both present
        if ex_val == 0:
            err_pct = float("inf") if sw_val != 0 else 0.0
        else:
            err_pct = abs(sw_val - ex_val) / abs(ex_val) * 100.0
        if err_pct <= tolerance_pct:
            matches.append((name, sw_val, ex_val, err_pct))
        else:
            mismatches.append((name, sw_val, ex_val, err_pct))
    return matches, sw_only, excel_only, mismatches


def render_report(sw_csv: Path, workbook: Path, sheet: str, config: str,
                  tolerance_pct: float, results: dict) -> str:
    matches = results["matches"]
    sw_only = results["sw_only"]
    excel_only = results["excel_only"]
    mismatches = results["mismatches"]
    lines = [
        f"# SolidWorks ↔ Excel Dimension Diff",
        "",
        f"- SW CSV: `{sw_csv}`",
        f"- Excel: `{workbook}` (sheet: `{sheet}`)",
        f"- Configuration: `{config or 'ALL'}`",
        f"- Tolerance: ±{tolerance_pct:.2f}%",
        f"- Generated: {datetime.now().isoformat(timespec='seconds')}",
        "",
        f"## Summary",
        "",
        f"| Bucket | Count |",
        f"|---|---|",
        f"| Matches | {len(matches)} |",
        f"| SW only | {len(sw_only)} |",
        f"| Excel only | {len(excel_only)} |",
        f"| **Mismatches** | **{len(mismatches)}** |",
        "",
    ]
    if mismatches:
        lines += [
            f"## Mismatches",
            "",
            f"| Variable | SW (in) | Excel (in) | Error % |",
            f"|---|---|---|---|",
        ]
        for name, sw_val, ex_val, err in sorted(
                mismatches, key=lambda x: -x[3]):
            lines.append(
                f"| `{name}` | {sw_val:.4f} | {ex_val:.4f} | {err:+.2f}% |"
            )
        lines.append("")
    if sw_only:
        lines += [f"## In SW but not in Excel", ""]
        for name, val in sorted(sw_only):
            lines.append(f"- `{name}` = {val:.4f} in")
        lines.append("")
    if excel_only:
        lines += [f"## In Excel but not in SW", ""]
        for name, val in sorted(excel_only):
            lines.append(f"- `{name}` = {val:.4f} in")
        lines.append("")
    if not (mismatches or sw_only or excel_only):
        lines += [
            f"## Result",
            "",
            f"All {len(matches)} dimensions match within tolerance. ✓",
            "",
        ]
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--csv", required=True, type=Path,
                    help="Path to Extract_Dimensions.swp CSV output")
    ap.add_argument("--workbook", required=True, type=Path,
                    help="Path to Excel design workbook")
    ap.add_argument("--design-sheet", default="Master_Inputs",
                    help="Sheet name in workbook (default: Master_Inputs)")
    ap.add_argument("--config", help="Restrict to one configuration")
    ap.add_argument("--tolerance-percent", type=float, default=0.5,
                    help="Tolerance in percent (default: 0.5)")
    ap.add_argument("--report", type=Path,
                    help="Write a Markdown report to this path")
    ap.add_argument("--write-validation", type=Path,
                    help="Append findings to a packet's validation.csv")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not args.csv.exists():
        print(f"CSV not found: {args.csv}", file=sys.stderr)
        return 1
    if not args.workbook.exists():
        print(f"workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    sw_by_config = read_sw_csv(args.csv, only_config=args.config)
    excel_dims = read_excel_master_inputs(args.workbook, args.design_sheet)

    print(f"SW configurations: {len(sw_by_config)}")
    print(f"Excel globals: {len(excel_dims)}")
    print()

    overall_results = {"matches": [], "sw_only": [], "excel_only": [],
                       "mismatches": []}
    for cfg, sw_dims in sw_by_config.items():
        m, so, eo, mm = diff(sw_dims, excel_dims, args.tolerance_percent)
        print(f"=== {cfg} ===")
        print(f"  matches: {len(m)}, sw_only: {len(so)}, "
              f"excel_only: {len(eo)}, mismatches: {len(mm)}")
        for name, sw_val, ex_val, err in mm:
            print(f"  ! {name}: SW={sw_val:.4f}in vs Excel={ex_val:.4f}in "
                  f"(err {err:+.2f}%)")
        # accumulate
        overall_results["matches"].extend(m)
        overall_results["sw_only"].extend(so)
        overall_results["excel_only"].extend(eo)
        overall_results["mismatches"].extend(mm)

    if args.report and not args.dry_run:
        # use the *first* config name (or 'ALL') as the diff context
        cfg_label = args.config or ("ALL" if len(sw_by_config) > 1
                                    else next(iter(sw_by_config), ""))
        report = render_report(args.csv, args.workbook, args.design_sheet,
                                cfg_label, args.tolerance_percent,
                                overall_results)
        args.report.write_text(report, encoding="utf-8")
        print(f"\nwrote report: {args.report}")

    if args.write_validation and not args.dry_run:
        with args.write_validation.open("a", newline="") as f:
            writer = csv.writer(f)
            for name, sw_val, ex_val, err in overall_results["mismatches"]:
                writer.writerow([
                    f"sw_diff_{name}", sw_val, ex_val,
                    f"±{args.tolerance_percent}%", "", "",
                    "Extract_Dimensions.swp",
                    "as-modeled vs Excel",
                    datetime.now().strftime("%Y-%m-%d"),
                    "fail" if err > args.tolerance_percent else "pass",
                    f"SW says {sw_val:.4f}in; Excel says {ex_val:.4f}in"
                    f" (err {err:+.2f}%)",
                ])
        print(f"\nappended {len(overall_results['mismatches'])} "
              f"finding(s) to {args.write_validation}")

    if overall_results["mismatches"]:
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
